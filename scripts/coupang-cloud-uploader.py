"""쿠팡 Excel을 Vercel 윤비서로 업로드하는 로컬 PC용 도구.

필수 환경변수:
  YUNBISEO_URL=https://내-윤비서.vercel.app
  YUNBISEO_API_KEY=윤비서 시스템설정에서 발급한 API 키

중요:
  판매분석 Excel은 조회기간을 반드시 하루로 설정한 파일이어야 합니다.
  쿠팡 아이디·비밀번호·로그인 쿠키는 전송하지 않습니다.
"""

import argparse
import hashlib
import json
import os
import socket
import urllib.error
import urllib.request
from datetime import date
from pathlib import Path

import openpyxl


def value(row, index, fallback=""):
    if index >= len(row) or row[index] is None:
        return fallback
    return row[index]


def integer(cell):
    if cell in (None, "", "-"):
        return 0
    text = str(cell).replace(",", "").replace("원", "").strip()
    try:
        return round(float(text))
    except ValueError:
        return 0


def read_sales(path: Path, sales_date: str):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in list(sheet.iter_rows(values_only=True))[1:]:
        option_id = str(value(row, 0)).strip()
        if not option_id:
            continue
        rows.append(
            {
                "sales_date": sales_date,
                "option_id": option_id,
                "option_name": str(value(row, 1)).strip(),
                "product_name": str(value(row, 2)).strip(),
                "seller_product_id": str(value(row, 3)).strip(),
                "sales_method": str(value(row, 5)).strip(),
                "gross_sales": integer(value(row, 6)),
                "order_quantity": integer(value(row, 7)),
                "sale_quantity": integer(value(row, 8)),
            }
        )
    workbook.close()
    return rows


def read_inventory(path: Path, snapshot_date: str):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in list(sheet.iter_rows(values_only=True))[2:]:
        option_id = str(value(row, 2)).strip()
        if not option_id:
            continue
        rows.append(
            {
                "snapshot_date": snapshot_date,
                "seller_product_id": str(value(row, 1)).strip(),
                "option_id": option_id,
                "sku": str(value(row, 3)).strip(),
                "product_name": str(value(row, 4)).strip(),
                "option_name": str(value(row, 5)).strip(),
                "channel_stock": integer(value(row, 7)),
                "inbound_quantity": integer(value(row, 8)),
                "warehouse_stock": 0,
                "reserved_quantity": 0,
            }
        )
    workbook.close()
    return rows


def post_payload(base_url: str, api_key: str, payload: dict):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        f"{base_url.rstrip('/')}/api/commerce/collector/upload",
        data=body,
        method="POST",
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "X-API-Key": api_key,
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"윤비서 업로드 실패 HTTP {error.code}: {detail}") from error


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--downloads", required=True, help="쿠팡 Excel 다운로드 폴더")
    parser.add_argument("--date", default=date.today().isoformat(), help="조회일 YYYY-MM-DD")
    parser.add_argument(
        "--account-type",
        default="wing_growth",
        choices=["wing_growth", "rocket"],
    )
    args = parser.parse_args()

    base_url = os.environ.get("YUNBISEO_URL", "").strip()
    api_key = os.environ.get("YUNBISEO_API_KEY", "").strip()
    if not base_url or not api_key:
        raise SystemExit("YUNBISEO_URL과 YUNBISEO_API_KEY 환경변수가 필요합니다.")

    download_dir = Path(args.downloads)
    sales_files = {p.name.removesuffix("_판매분석.xlsx"): p for p in download_dir.glob("*_판매분석.xlsx")}
    inventory_files = {p.name.removesuffix("_재고현황.xlsx"): p for p in download_dir.glob("*_재고현황.xlsx")}
    accounts = sorted(set(sales_files) | set(inventory_files))
    if not accounts:
        raise SystemExit("업로드할 쿠팡 Excel 파일이 없습니다.")

    for account_name in accounts:
        sales = read_sales(sales_files[account_name], args.date) if account_name in sales_files else []
        inventory = (
            read_inventory(inventory_files[account_name], args.date)
            if account_name in inventory_files
            else []
        )
        fingerprint = hashlib.sha256(
            json.dumps(
                [account_name, args.date, sales, inventory],
                ensure_ascii=False,
                sort_keys=True,
            ).encode("utf-8")
        ).hexdigest()
        payload = {
            "batch_key": f"coupang:{account_name}:{args.date}:{fingerprint}",
            "device_name": socket.gethostname(),
            "account_key": account_name,
            "account_name": account_name,
            "account_type": args.account_type,
            "sales": sales,
            "inventory": inventory,
        }
        result = post_payload(base_url, api_key, payload)
        print(f"[{account_name}] {result.get('message', result)}")


if __name__ == "__main__":
    main()
